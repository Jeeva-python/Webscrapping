import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from datetime import datetime

# Load data from output.txt with correct header row
file_path = 'output.txt'

# Read the data with a single header
data = pd.read_csv(file_path, sep='\t', dtype=str, header=0)  # Use the first row as header

# Remove duplicate headers if they exist within the file data
data = data.loc[~data.apply(lambda row: row.str.contains('Strike ID').any(), axis=1)]

# Remove duplicates based on 'Strike ID'
data = data.drop_duplicates(subset='Strike ID')

# Sort data by 'Strike ID'
data = data.sort_values(by='Strike ID')

# Define the file name with current date
current_date = datetime.now().strftime("%m.%d.%Y")
excel_filename = f'ReportfromStrikeaprice_Techforless_Serversupply.com_{current_date}.xlsx'

# Save to Excel
data.to_excel(excel_filename, sheet_name='Report', index=False)

# Load the workbook and sheet
wb = load_workbook(excel_filename)
ws = wb['Report']

# Apply font, border, and header fill formatting
font = Font(name='Cambria', size=10)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # RGB color: 0, 176, 80

# Apply font, border, and alignment formatting
for row in ws.iter_rows():
    for cell in row:
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal='left')  # Align all cells to the left

# Apply header formatting
for cell in ws[1]:
    cell.fill = header_fill

# Set column widths to 13.5
column_width = 13.5
for col in ws.columns:
    column_letter = col[0].column_letter
    ws.column_dimensions[column_letter].width = column_width

# Set tab color to RGB 0, 176, 80
ws.sheet_properties.tabColor = "00B050"  # Hex color code for RGB 0, 176, 80

# Save the workbook
wb.save(excel_filename)
